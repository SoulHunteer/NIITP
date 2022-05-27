import sys  # sys нужен для передачи argv в QApplication
from PyQt5 import QtWidgets
import mainwindow
from openpyxl import load_workbook, Workbook
import datetime
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *


class ExampleApp(QtWidgets.QMainWindow, mainwindow.Ui_MainWindow):
    # В КОДЕ НИЖЕ УКАЗЫВАЮТСЯ ЗНАЧЕНИЯ ПО УМОЛЧАНИЮ, КОТОРЫЕ ПРИ ЖЕЛАНИИ МОЖНО ИЗМЕНИТЬ, КОММЕНТЫ Я ОСТАВИЛ
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.pushButton.clicked.connect(self.action)
        self.lineEdit.setText(
            r'C:\Users\Sergey\PycharmProjects\NIITP\Электронный_журнал_статистики_ЕТРИС_ДЗЗ_2022.xlsx')  # Здесь выбирается дефолтный файл Excel
        self.lineEdit_2.setText(r'НКПОР-Р-В (Восточный)')  # Здесь выбирается дефолтный лист Excel
        self.dateEdit.setDate(
            QDate(QDate.year(QDate.currentDate()), 1, 1))  # Тут выставляется дата, в одну строчку не опишешь как :)
        self.dateEdit_2.setDate(QDate.currentDate())  # Тут выставляется дата, в одну строчку не опишешь как :)
        self.spinBox.setValue(7)  # Тут выставляется значение по умолчанию для колонки с ошибками
        self.spinBox_2.setValue(12)  # Тут выставляется значение по умолчанию для колонки с комментариями

    def action(self, path):
        self.book = self.lineEdit_2.text()
        self.path = self.lineEdit.text()
        self.from_date = self.dateEdit.date().toPyDate()
        self.to_date = self.dateEdit_2.date().toPyDate()
        self.column_error = self.spinBox.text()
        self.column_comm = self.spinBox_2.text()
        try:
            book = load_workbook(
                self.path.strip())
            sheet = book[self.book]
            result_list = []
            for i, row in enumerate(sheet.rows):
                if i == 0:
                    continue
                if row[0].value is not None and row[int(self.column_comm)].value is not None and row[int(self.column_error)].value != 0 and \
                        row[0].value.date() >= self.from_date and row[0].value.date() <= self.to_date:
                    result_list.append([row[0].value, row[int(self.column_comm)].value])

            workbook = Workbook()
            current_time = str(datetime.datetime.now()).replace(' ', '_').replace(':', '_').replace('.', '_')
            name = current_time + '_' + 'отчёт_об_ошибках_' + self.book + '.xlsx'
            workbook.save(current_time + '_' + 'отчёт_об_ошибках_' + self.book + '.xlsx')
            new_book = load_workbook(current_time + '_' + 'отчёт_об_ошибках_' + self.book + '.xlsx')
            new_sheet = new_book.active
            for row in result_list:
                new_sheet.append(row)
            new_book.save(current_time + '_' + 'отчёт_об_ошибках_' + self.book + '.xlsx')
            QMessageBox.about(self, "Успешно", f"Файл {name} сформирован успешно!")
        except Exception as e:
            QMessageBox.about(self, "Ошибка!", e.__str__())
            print(e)


def main():
    app = QtWidgets.QApplication(sys.argv)
    window = ExampleApp()
    window.show()
    app.exec_()


if __name__ == '__main__':
    main()
